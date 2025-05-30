"""
Manifest export service for generating formatted Excel and CSV files.
"""
import logging
import pandas as pd
from io import BytesIO, StringIO
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.utils import get_column_letter

# Set up logger for this module
logger = logging.getLogger(__name__)

class ManifestExportService:
    """
    Service for exporting manifest data to various formats with enhanced formatting.
    """

    @classmethod
    def export_remapped_manifest(cls, manifest, items, format='xlsx'):
        """
        Export manifest data in the specified format with enhanced formatting.
        
        Args:
            manifest: The Manifest model instance
            items: QuerySet of ManifestItems
            format: Output format ('xlsx' or 'csv')
            
        Returns:
            HttpResponse with the appropriate file
        """
        try:
            # Create a DataFrame with the mapped data
            data = []
            for item in items:
                # Extract all mapped fields
                item_data = {
                    'Serial Number': item.serial,
                    'Manufacturer': item.manufacturer,
                    'Model': item.model,
                    'Processor': item.processor,
                    'Memory': item.memory,
                    'Storage': item.storage,
                    'Condition Grade': item.condition_grade,
                    'Barcode': item.barcode,
                    # Add other fields as needed
                }
                data.append(item_data)
            
            df = pd.DataFrame(data)
            
            # Create a response with the right content type
            if format == 'xlsx':
                # Generate Excel file with enhanced formatting
                return cls._generate_excel(manifest, df)
            else:  # CSV format
                # Generate CSV with a hidden signature
                return cls._generate_csv(manifest, df)
            
        except Exception as e:
            logger.error(f"Error in export_remapped_manifest: {str(e)}", exc_info=True)
            raise

    @classmethod
    def _generate_excel(cls, manifest, df):
        """
        Generate a well-formatted Excel file with data, summary, and hidden signature.
        
        Args:
            manifest: The Manifest model instance
            df: DataFrame with the manifest data
            
        Returns:
            HttpResponse with the Excel file
        """
        try:
            # Create BytesIO buffer to store Excel file
            buffer = BytesIO()
            
            # Create Excel writer with xlsxwriter engine for better formatting
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                # Write main data to Data sheet
                df.to_excel(writer, index=False, sheet_name='Data')
                
                # Access the workbook and sheets
                workbook = writer.book
                data_sheet = workbook['Data']
                
                # Apply formatting to the data sheet
                cls._format_data_sheet(data_sheet, df)
                
                # Generate and add a summary sheet
                cls._add_summary_sheet(workbook, df, manifest)
                
                # Add hidden signature sheet
                cls._add_signature_sheet(workbook, manifest)
                
                # Set document properties
                cls._set_document_properties(workbook, manifest)
            
            # Set response headers
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=manifest_{manifest.id}_remapped.xlsx'
            logger.info("Successfully created Excel response")
            return response
            
        except Exception as e:
            logger.error(f"Error generating Excel: {str(e)}", exc_info=True)
            raise
    
    @staticmethod
    def _format_data_sheet(worksheet, df):
        """
        Apply professional formatting to the data worksheet.
        
        Args:
            worksheet: The openpyxl worksheet
            df: The pandas DataFrame
        """
        try:
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
            from openpyxl.utils import get_column_letter
            
            # Create named styles for reuse
            header_style = NamedStyle(name="header_style")
            header_style.font = Font(bold=True, color="FFFFFF", size=11)
            header_style.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            header_style.border = Border(
                left=Side(style="thin"), 
                right=Side(style="thin"), 
                top=Side(style="thin"), 
                bottom=Side(style="thin")
            )
            header_style.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Add style to workbook if it doesn't exist
            if header_style.name not in worksheet.parent.named_styles:
                worksheet.parent.add_named_style(header_style)
            
            # Apply header styles
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.style = header_style
                
                # Make header bold and centered
                cell.font = header_style.font
                cell.fill = header_style.fill
                cell.border = header_style.border
                cell.alignment = header_style.alignment
            
            # Auto-adjust columns width with better algorithm
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter  # Get column name
                
                for cell in col:
                    if cell.row == 1:  # Header row
                        max_length = max(max_length, len(str(cell.value)) + 4)  # Headers need more space
                    else:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value)) + 2
                        except:
                            pass
                
                # Set adjusted width with min/max constraints
                adjusted_width = max(max_length, 12)  # Minimum width of 12
                worksheet.column_dimensions[column].width = min(adjusted_width, 40)  # Maximum width of 40
                
            # Add zebra striping to data rows with a softer color
            for row in range(2, worksheet.max_row + 1):
                if row % 2 == 0:
                    light_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                    for col in range(1, worksheet.max_column + 1):
                        worksheet.cell(row=row, column=col).fill = light_fill
            
            # Add borders to all cells with a softer border color
            border = Border(
                left=Side(style="thin", color="D3D3D3"), 
                right=Side(style="thin", color="D3D3D3"), 
                top=Side(style="thin", color="D3D3D3"), 
                bottom=Side(style="thin", color="D3D3D3")
            )
            
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    if cell.row > 1:  # Skip header row which already has its own style
                        cell.border = border
                        cell.alignment = Alignment(vertical="center")  # Vertically center all data
            
            # Freeze the header row
            worksheet.freeze_panes = "A2"
            
            # Add a filter to the header row
            worksheet.auto_filter.ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
            
        except Exception as e:
            logger.error(f"Error formatting data sheet: {str(e)}", exc_info=True)
    
    @staticmethod
    def _add_summary_sheet(workbook, df, manifest):
        """
        Add a summary sheet with statistics, charts and improved styling.
        
        Args:
            workbook: The openpyxl workbook
            df: The pandas DataFrame
            manifest: The Manifest model instance
        """
        try:
            # Create summary sheet
            summary_sheet = workbook.create_sheet("Summary", 0)  # Add as the first sheet
            
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.chart import BarChart, PieChart, Reference
            from openpyxl.drawing.image import Image
            from openpyxl.worksheet.dimensions import RowDimension, ColumnDimension
            
            # Define styles
            title_font = Font(name='Calibri', size=16, bold=True, color="000080")  # Dark blue title
            subtitle_font = Font(name='Calibri', size=13, bold=True, color="0066CC")
            header_font = Font(name='Calibri', size=11, bold=True)
            
            title_fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")
            header_fill = PatternFill(start_color="EDF2F7", end_color="EDF2F7", fill_type="solid")
            section_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
            
            border = Border(
                left=Side(style="thin", color="D3D3D3"), 
                right=Side(style="thin", color="D3D3D3"), 
                top=Side(style="thin", color="D3D3D3"), 
                bottom=Side(style="thin", color="D3D3D3")
            )
            
            thick_border = Border(
                bottom=Side(style="medium", color="000080")
            )
            
            # Set column widths for better spacing
            columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
            for col in columns:
                summary_sheet.column_dimensions[col].width = 18
            
            # REPORT HEADER SECTION
            # --------------------
            # Set row heights for better spacing in header section
            summary_sheet.row_dimensions[1].height = 30
            summary_sheet.row_dimensions[2].height = 20
            
            # Add title and metadata with improved styling
            main_title_cell = summary_sheet['A1'] 
            main_title_cell.value = f"Manifest Summary Report"
            main_title_cell.font = title_font
            main_title_cell.alignment = Alignment(horizontal="left", vertical="center")
            summary_sheet.merge_cells('A1:H1')
            
            subtitle_cell = summary_sheet['A2']
            subtitle_cell.value = f"{manifest.name}"
            subtitle_cell.font = subtitle_font
            subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")
            summary_sheet.merge_cells('A2:H2')
            
            # Add a separator line
            for col in range(1, 9):  # A through H
                cell = summary_sheet.cell(row=3, column=col)
                cell.border = thick_border
            
            # MANIFEST INFORMATION SECTION
            # ---------------------------
            # Add manifest information with improved styling
            summary_sheet['A5'] = "Manifest Information"
            summary_sheet['A5'].font = subtitle_font
            summary_sheet['A5'].fill = section_fill
            summary_sheet['A5'].alignment = Alignment(horizontal="left", vertical="center")
            summary_sheet.merge_cells('A5:D5')
            summary_sheet.row_dimensions[5].height = 25
            
            # Create two columns of information for better layout
            info_rows_col1 = [
                ('Manifest ID:', f"{manifest.id}"),
                ('Status:', f"{manifest.status.upper() if manifest.status else 'N/A'}"),
                ('Total Items:', f"{len(df)}"),
            ]
            
            info_rows_col2 = [
                ('Date Created:', f"{manifest.uploaded_at.strftime('%B %d, %Y') if manifest.uploaded_at else 'N/A'}"),
                ('Reference:', f"{manifest.reference or 'N/A'}"),
                ('Export Date:', f"{timezone.now().strftime('%B %d, %Y at %H:%M')}"),
            ]
            
            # Apply styling to info rows - first column
            for i, (label, value) in enumerate(info_rows_col1):
                row = 6 + i
                summary_sheet[f'A{row}'] = label
                summary_sheet[f'A{row}'].font = header_font
                summary_sheet[f'A{row}'].alignment = Alignment(horizontal="right", vertical="center")
                
                summary_sheet[f'B{row}'] = value
                summary_sheet[f'B{row}'].alignment = Alignment(vertical="center")
                
                # Add subtle borders
                summary_sheet[f'A{row}'].border = border
                summary_sheet[f'B{row}'].border = border
            
            # Apply styling to info rows - second column
            for i, (label, value) in enumerate(info_rows_col2):
                row = 6 + i
                summary_sheet[f'C{row}'] = label
                summary_sheet[f'C{row}'].font = header_font
                summary_sheet[f'C{row}'].alignment = Alignment(horizontal="right", vertical="center")
                
                summary_sheet[f'D{row}'] = value
                summary_sheet[f'D{row}'].alignment = Alignment(vertical="center")
                
                # Add subtle borders
                summary_sheet[f'C{row}'].border = border
                summary_sheet[f'D{row}'].border = border
            
            # Add spacing before statistics section
            summary_sheet.row_dimensions[10].height = 15
            
            # STATISTICS SECTION
            # -----------------
            section_row = 11
            summary_sheet[f'A{section_row}'] = "Inventory Statistics & Distribution"
            summary_sheet[f'A{section_row}'].font = subtitle_font
            summary_sheet[f'A{section_row}'].fill = section_fill
            summary_sheet[f'A{section_row}'].alignment = Alignment(horizontal="left", vertical="center")
            summary_sheet.merge_cells(f'A{section_row}:H{section_row}')
            summary_sheet.row_dimensions[section_row].height = 25
            
            # MANUFACTURER STATS
            # -----------------
            row = section_row + 2
            summary_sheet[f'A{row}'] = "Distribution by Manufacturer"
            summary_sheet[f'A{row}'].font = subtitle_font
            summary_sheet[f'A{row}'].alignment = Alignment(horizontal="center", vertical="center")
            summary_sheet.merge_cells(f'A{row}:B{row}')
            summary_sheet.row_dimensions[row].height = 20
            
            # Calculate manufacturer counts
            if 'Manufacturer' in df.columns:
                manufacturer_counts = df['Manufacturer'].value_counts().reset_index()
                manufacturer_counts.columns = ['Manufacturer', 'Count']
                
                # Add percentage column
                manufacturer_counts['Percentage'] = (manufacturer_counts['Count'] / manufacturer_counts['Count'].sum() * 100).round(1)
                manufacturer_counts['Percentage'] = manufacturer_counts['Percentage'].apply(lambda x: f"{x}%")
                
                # Write manufacturer data with headers
                header_row = row + 1
                summary_sheet[f'A{header_row}'] = "Manufacturer"
                summary_sheet[f'A{header_row}'].font = header_font
                summary_sheet[f'A{header_row}'].fill = header_fill
                summary_sheet[f'A{header_row}'].alignment = Alignment(horizontal="center", vertical="center")
                summary_sheet[f'A{header_row}'].border = border
                
                summary_sheet[f'B{header_row}'] = "Count"
                summary_sheet[f'B{header_row}'].font = header_font
                summary_sheet[f'B{header_row}'].fill = header_fill
                summary_sheet[f'B{header_row}'].alignment = Alignment(horizontal="center", vertical="center")
                summary_sheet[f'B{header_row}'].border = border
                
                summary_sheet[f'C{header_row}'] = "Percentage"
                summary_sheet[f'C{header_row}'].font = header_font
                summary_sheet[f'C{header_row}'].fill = header_fill
                summary_sheet[f'C{header_row}'].alignment = Alignment(horizontal="center", vertical="center")
                summary_sheet[f'C{header_row}'].border = border
                
                # Write manufacturer data rows
                for i, (_, row_data) in enumerate(manufacturer_counts.iterrows()):
                    r = header_row + 1 + i
                    
                    summary_sheet[f'A{r}'] = row_data['Manufacturer'] or "Unknown"
                    summary_sheet[f'A{r}'].border = border
                    
                    summary_sheet[f'B{r}'] = row_data['Count']
                    summary_sheet[f'B{r}'].border = border
                    
                    summary_sheet[f'C{r}'] = row_data['Percentage']
                    summary_sheet[f'C{r}'].border = border
                
                # Create a pie chart for manufacturers
                pie = PieChart()
                labels = Reference(summary_sheet, min_col=1, min_row=header_row+1, max_row=header_row+len(manufacturer_counts))
                data = Reference(summary_sheet, min_col=2, min_row=header_row, max_row=header_row+len(manufacturer_counts))
                pie.add_data(data, titles_from_data=True)
                pie.set_categories(labels)
                pie.title = "Manufacturers Distribution"
                pie.style = 10  # Use a nicer style
                
                # Make the chart larger and add a legend
                pie.height = 15
                pie.width = 15
                pie.legend.position = 'r'
                
                # Place chart in summary sheet
                chart_col = get_column_letter(5)  # Column E
                chart_row = row + 1
                summary_sheet.add_chart(pie, f"{chart_col}{chart_row}")
                
                # MODEL STATS
                # -----------
                model_row = row + manufacturer_counts.shape[0] + 5
                
                summary_sheet[f'A{model_row}'] = "Distribution by Model"
                summary_sheet[f'A{model_row}'].font = subtitle_font
                summary_sheet[f'A{model_row}'].alignment = Alignment(horizontal="center", vertical="center")
                summary_sheet.merge_cells(f'A{model_row}:B{model_row}')
                summary_sheet.row_dimensions[model_row].height = 20
            
            # Calculate model counts
            if 'Model' in df.columns:
                model_counts = df['Model'].value_counts().head(10).reset_index()  # Top 10 models
                model_counts.columns = ['Model', 'Count']
                
                # Add percentage column
                model_counts['Percentage'] = (model_counts['Count'] / df.shape[0] * 100).round(1)
                model_counts['Percentage'] = model_counts['Percentage'].apply(lambda x: f"{x}%")
                
                # Write model data with headers
                model_header_row = model_row + 1
                summary_sheet[f'A{model_header_row}'] = "Model"
                summary_sheet[f'A{model_header_row}'].font = header_font
                summary_sheet[f'A{model_header_row}'].fill = header_fill
                summary_sheet[f'A{model_header_row}'].alignment = Alignment(horizontal="center", vertical="center")
                summary_sheet[f'A{model_header_row}'].border = border
                
                summary_sheet[f'B{model_header_row}'] = "Count"
                summary_sheet[f'B{model_header_row}'].font = header_font
                summary_sheet[f'B{model_header_row}'].fill = header_fill
                summary_sheet[f'B{model_header_row}'].alignment = Alignment(horizontal="center", vertical="center")
                summary_sheet[f'B{model_header_row}'].border = border
                
                summary_sheet[f'C{model_header_row}'] = "Percentage"
                summary_sheet[f'C{model_header_row}'].font = header_font
                summary_sheet[f'C{model_header_row}'].fill = header_fill
                summary_sheet[f'C{model_header_row}'].alignment = Alignment(horizontal="center", vertical="center")
                summary_sheet[f'C{model_header_row}'].border = border
                
                # Write model data
                for i, (_, row_data) in enumerate(model_counts.iterrows()):
                    r = model_header_row + 1 + i
                    
                    summary_sheet[f'A{r}'] = row_data['Model'] or "Unknown"
                    summary_sheet[f'A{r}'].border = border
                    
                    summary_sheet[f'B{r}'] = row_data['Count']
                    summary_sheet[f'B{r}'].border = border
                    
                    summary_sheet[f'C{r}'] = row_data['Percentage']
                    summary_sheet[f'C{r}'].border = border
                
                # Create a horizontal bar chart for models (better for text labels)
                chart = BarChart()
                chart.type = "bar"  # Horizontal bar chart
                chart.style = 12
                chart.title = "Top Models by Count"
                chart.y_axis.title = "Model"
                chart.x_axis.title = "Count"
                chart.height = 15
                chart.width = 20
                
                data = Reference(summary_sheet, min_col=2, min_row=model_header_row, 
                                max_row=model_header_row+len(model_counts))
                cats = Reference(summary_sheet, min_col=1, min_row=model_header_row+1, 
                                max_row=model_header_row+len(model_counts))
                
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                
                # Place chart in summary sheet
                chart_col = get_column_letter(5)  # Column E
                chart_row = model_row + 1
                summary_sheet.add_chart(chart, f"{chart_col}{chart_row}")
                
                # Add a note if showing only top models
                if len(df['Model'].value_counts()) > 10:
                    note_row = model_header_row + len(model_counts) + 1
                    summary_sheet[f'A{note_row}'] = "* Showing top 10 models only"
                    summary_sheet[f'A{note_row}'].font = Font(italic=True, size=8)
                    summary_sheet.merge_cells(f'A{note_row}:C{note_row}')
            
            # CONDITION GRADE STATS
            # --------------------
            if 'Condition Grade' in df.columns:
                grade_row = model_row + len(model_counts) + 7 if 'Model' in df.columns else model_row + 7
                
                summary_sheet[f'A{grade_row}'] = "Distribution by Condition Grade"
                summary_sheet[f'A{grade_row}'].font = subtitle_font
                summary_sheet[f'A{grade_row}'].alignment = Alignment(horizontal="center", vertical="center")
                summary_sheet.merge_cells(f'A{grade_row}:C{grade_row}')
                summary_sheet.row_dimensions[grade_row].height = 20
                
                # Calculate grade counts
                grade_counts = df['Condition Grade'].value_counts().reset_index()
                grade_counts.columns = ['Grade', 'Count']
                
                # Add percentage column
                grade_counts['Percentage'] = (grade_counts['Count'] / grade_counts['Count'].sum() * 100).round(1)
                grade_counts['Percentage'] = grade_counts['Percentage'].apply(lambda x: f"{x}%")
                
                # Write grade data with headers
                grade_header_row = grade_row + 1
                summary_sheet[f'A{grade_header_row}'] = "Grade"
                summary_sheet[f'A{grade_header_row}'].font = header_font
                summary_sheet[f'A{grade_header_row}'].fill = header_fill
                summary_sheet[f'A{grade_header_row}'].alignment = Alignment(horizontal="center", vertical="center")
                summary_sheet[f'A{grade_header_row}'].border = border
                
                summary_sheet[f'B{grade_header_row}'] = "Count"
                summary_sheet[f'B{grade_header_row}'].font = header_font
                summary_sheet[f'B{grade_header_row}'].fill = header_fill
                summary_sheet[f'B{grade_header_row}'].alignment = Alignment(horizontal="center", vertical="center")
                summary_sheet[f'B{grade_header_row}'].border = border
                
                summary_sheet[f'C{grade_header_row}'] = "Percentage"
                summary_sheet[f'C{grade_header_row}'].font = header_font
                summary_sheet[f'C{grade_header_row}'].fill = header_fill
                summary_sheet[f'C{grade_header_row}'].alignment = Alignment(horizontal="center", vertical="center")
                summary_sheet[f'C{grade_header_row}'].border = border
                
                # Write grade data
                for i, (_, row_data) in enumerate(grade_counts.iterrows()):
                    r = grade_header_row + 1 + i
                    
                    summary_sheet[f'A{r}'] = row_data['Grade'] or "Unknown"
                    summary_sheet[f'A{r}'].border = border
                    
                    summary_sheet[f'B{r}'] = row_data['Count']
                    summary_sheet[f'B{r}'].border = border
                    
                    summary_sheet[f'C{r}'] = row_data['Percentage']
                    summary_sheet[f'C{r}'].border = border
                
                # Create a bar chart for condition grades
                chart = BarChart()
                chart.type = "col"
                chart.style = 10
                chart.title = "Condition Grade Distribution"
                chart.y_axis.title = "Count"
                chart.x_axis.title = "Grade"
                chart.height = 15
                chart.width = 10
                
                data = Reference(summary_sheet, min_col=2, min_row=grade_header_row, 
                                max_row=grade_header_row+len(grade_counts))
                cats = Reference(summary_sheet, min_col=1, min_row=grade_header_row+1, 
                                max_row=grade_header_row+len(grade_counts))
                
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                
                # Place chart in summary sheet
                chart_col = get_column_letter(5)  # Column E
                chart_row = grade_row + 1
                summary_sheet.add_chart(chart, f"{chart_col}{chart_row}")
                
            # Add a footer with the export date at the bottom
            footer_row = grade_row + 20 if 'Condition Grade' in df.columns else grade_row + 10
            summary_sheet[f'A{footer_row}'] = f"Report generated on {timezone.now().strftime('%Y-%m-%d at %H:%M:%S')}"
            summary_sheet[f'A{footer_row}'].font = Font(italic=True, size=8, color="6C757D")
            summary_sheet.merge_cells(f'A{footer_row}:D{footer_row}')
            
        except Exception as e:
            logger.error(f"Error creating summary sheet: {str(e)}", exc_info=True)
            # If there's an error with the summary sheet, we'll just continue without it
            if "Summary" in workbook.sheetnames:
                workbook.remove(workbook["Summary"])
            logger.info("Continuing without summary sheet due to error")
    
    @staticmethod
    def _add_signature_sheet(workbook, manifest):
        """
        Add a hidden signature sheet with company information.
        
        Args:
            workbook: The openpyxl workbook
            manifest: The Manifest model instance
        """
        # Add hidden worksheet with company signature
        signature_sheet = workbook.create_sheet("_signature", -1)  # Add at the end
        signature_sheet.sheet_state = 'hidden'
        
        # Add company signature information with better formatting
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        title_font = Font(name='Calibri', size=12, bold=True, color="000080")
        normal_font = Font(name='Calibri', size=10)
        
        signature_sheet['A1'] = "REPLUGIT DATA EXPORT"
        signature_sheet['A1'].font = title_font
        
        data_rows = [
            ("Manifest ID:", manifest.id),
            ("Manifest Name:", manifest.name),
            ("Export Date:", timezone.now().strftime('%Y-%m-%d %H:%M:%S')),
            ("User:", 'Anonymous' if not hasattr(manifest, 'uploaded_by') or not manifest.uploaded_by else manifest.uploaded_by.username),
            ("System:", "Replugit Manifest Management System"),
        ]
        
        for i, (label, value) in enumerate(data_rows):
            row = i + 2
            signature_sheet[f'A{row}'] = label
            signature_sheet[f'A{row}'].font = Font(bold=True)
            
            signature_sheet[f'B{row}'] = value
            signature_sheet[f'B{row}'].font = normal_font
        
        signature_sheet['A8'] = "This file contains proprietary data from Replugit."
        signature_sheet['A8'].font = Font(italic=True, size=8)
        
        # Add checksum for data integrity
        import hashlib
        checksum = hashlib.md5(f"{manifest.id}_{timezone.now()}".encode()).hexdigest()
        signature_sheet['A10'] = "File checksum:"
        signature_sheet['B10'] = checksum
    
    @staticmethod
    def _set_document_properties(workbook, manifest):
        """
        Set document metadata properties.
        
        Args:
            workbook: The openpyxl workbook
            manifest: The Manifest model instance
        """
        # Add document properties
        workbook.properties.creator = "Replugit Data Export System"
        workbook.properties.title = f"Manifest {manifest.name} - Replugit Export"
        workbook.properties.description = f"Exported from Replugit on {timezone.now().strftime('%Y-%m-%d')}"
        workbook.properties.keywords = f"replugit,manifest,{manifest.id}"
        workbook.properties.category = "Manifest Export"
        workbook.properties.lastModifiedBy = "Replugit System"
    
    @classmethod
    def _generate_csv(cls, manifest, df):
        """
        Generate a CSV file with the data and a hidden signature.
        
        Args:
            manifest: The Manifest model instance
            df: DataFrame with the manifest data
            
        Returns:
            HttpResponse with the CSV file
        """
        # Create CSV with a hidden signature line at the end
        buffer = StringIO()
        
        # Write data to CSV
        df.to_csv(buffer, index=False)
        
        # Add hidden signature as a comment line
        signature = f"# REPLUGIT DATA EXPORT - Manifest ID: {manifest.id} - {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
        buffer.write(f"\n{signature}")
        
        # Set response headers
        response = HttpResponse(buffer.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename=manifest_{manifest.id}_remapped.csv'
        logger.info("Successfully created CSV response")
        return response